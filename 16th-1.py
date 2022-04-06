from openpyxl import load_workbook

wb = load_workbook('result.xlsx')
ws = wb.active

print(ws['A1'].value)